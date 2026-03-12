"""Parse Moodle XLSX enrolment exports into validated enrolment entries."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from promptgrimoire.validation import is_valid_email

_REQUIRED_COLUMNS = ("first name", "last name", "id number", "email address")
_OPTIONAL_COLUMNS = ("groups",)


@dataclass(frozen=True)
class EnrolmentEntry:
    """A single validated student enrolment row."""

    email: str
    display_name: str
    student_id: str
    groups: tuple[str, ...]


class EnrolmentParseError(Exception):
    """Raised on validation failure. Carries line-numbered error strings."""

    def __init__(self, errors: list[str]) -> None:
        self.errors = errors
        super().__init__(f"Enrolment parse errors: {'; '.join(errors)}")


def _parse_groups(raw: str | None) -> tuple[str, ...]:
    """Parse ``[Tutorial 1], [Lab A]`` into a tuple of group names."""
    if not raw:
        return ()
    parts: list[str] = []
    for chunk in raw.split(","):
        cleaned = chunk.strip().strip("[]").strip()
        if cleaned:
            parts.append(cleaned)
    return tuple(parts)


def _resolve_columns(
    header_row: tuple[object, ...],
) -> tuple[dict[str, int], list[str]]:
    """Map normalised column names to 0-based indexes.

    Returns (column_map, missing) where *missing* lists required columns
    not found in the header row.
    """
    normalised = {
        str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell
    }
    col_map: dict[str, int] = {}
    missing: list[str] = []
    for name in _REQUIRED_COLUMNS:
        if name in normalised:
            col_map[name] = normalised[name]
        else:
            missing.append(name)
    for name in _OPTIONAL_COLUMNS:
        if name in normalised:
            col_map[name] = normalised[name]
    return col_map, missing


def _is_blank_row(row: tuple, col_map: dict[str, int]) -> bool:
    """Return True when every mapped cell is None or whitespace-only."""
    for idx in col_map.values():
        val = row[idx] if idx < len(row) else None
        if val is not None and str(val).strip():
            return False
    return True


def _cell_str(row: tuple, idx: int) -> str:
    """Extract a cell value as a stripped string, defaulting to ``""``."""
    val = row[idx] if idx < len(row) else None
    if val is None:
        return ""
    return str(val).strip()


def parse_xlsx(data: bytes) -> list[EnrolmentEntry]:
    """Parse XLSX bytes into a list of validated enrolment entries.

    Raises :class:`EnrolmentParseError` if any validation failures are found.
    All errors are collected before raising (not fail-fast).
    """
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    col_map, missing = _resolve_columns(rows[0])
    if missing:
        raise EnrolmentParseError(
            [f"Missing required column: {name}" for name in missing]
        )

    errors: list[str] = []
    entries: list[EnrolmentEntry] = []
    seen_emails: dict[str, int] = {}  # lowercased email -> first XLSX row number

    for row_idx, row in enumerate(rows[1:], start=2):
        if _is_blank_row(row, col_map):
            continue

        first_name = _cell_str(row, col_map["first name"])
        last_name = _cell_str(row, col_map["last name"])
        student_id = _cell_str(row, col_map["id number"])
        email = _cell_str(row, col_map["email address"])

        # Email validation
        if not is_valid_email(email):
            errors.append(f"Row {row_idx}: invalid email '{email}'")
            continue

        # Duplicate detection (case-insensitive)
        email_lower = email.lower()
        if email_lower in seen_emails:
            first_row = seen_emails[email_lower]
            errors.append(
                f"Duplicate email '{email}' at rows {first_row} and {row_idx}"
            )
            continue
        seen_emails[email_lower] = row_idx

        # Groups (optional column)
        groups_raw = _cell_str(row, col_map["groups"]) if "groups" in col_map else None
        groups = _parse_groups(groups_raw)

        entries.append(
            EnrolmentEntry(
                email=email,
                display_name=f"{first_name} {last_name}",
                student_id=student_id,
                groups=groups,
            )
        )

    if errors:
        raise EnrolmentParseError(errors)

    return entries
